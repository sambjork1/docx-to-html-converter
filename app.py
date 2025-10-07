from flask import Flask, render_template, request, send_file
from werkzeug.utils import secure_filename
import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import unicodedata
import re
from docx import Document

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # Max 50MB total

def slugify(text):
    text = unicodedata.normalize('NFKD', text)
    text = text.encode('ascii', 'ignore').decode('ascii')
    text = re.sub(r'[^\w\s-]', '', text.lower())
    text = re.sub(r'[-\s]+', '-', text)
    return text.strip('-')

def get_numbering_format(para, doc):
    """Hämtar numreringsformatet från Word's numbering.xml"""
    try:
        if para._element.pPr is None:
            return None
        
        numPr = para._element.pPr.numPr
        if numPr is None:
            return None
        
        numId_element = numPr.numId
        if numId_element is None:
            return None
        
        numId = numId_element.val
        ilvl_element = numPr.ilvl
        ilvl = ilvl_element.val if ilvl_element is not None else 0
        
        numbering_part = doc.part.numbering_part
        if numbering_part is None:
            return 'bullet'
        
        num_element = None
        for num in numbering_part.element.findall('.//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}num'):
            if num.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}numId') == str(numId):
                num_element = num
                break
        
        if num_element is None:
            return 'bullet'
        
        abstractNumId_element = num_element.find('.//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}abstractNumId')
        if abstractNumId_element is None:
            return 'bullet'
        
        abstractNumId = abstractNumId_element.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val')
        
        abstractNum = None
        for anum in numbering_part.element.findall('.//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}abstractNum'):
            if anum.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}abstractNumId') == str(abstractNumId):
                abstractNum = anum
                break
        
        if abstractNum is None:
            return 'bullet'
        
        for lvl in abstractNum.findall('.//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}lvl'):
            if lvl.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}ilvl') == str(ilvl):
                numFmt = lvl.find('.//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}numFmt')
                if numFmt is not None:
                    fmt_val = numFmt.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val')
                    if fmt_val in ['decimal', 'lowerLetter', 'upperLetter', 'lowerRoman', 'upperRoman']:
                        return 'numbered'
                    else:
                        return 'bullet'
        
        return 'bullet'
    except Exception as e:
        print(f"Error in get_numbering_format: {e}")
        return 'bullet'

def parse_docx(filepath):
    doc = Document(filepath)
    
    meta_title = ""
    meta_description = ""
    
    i = 0
    paras = doc.paragraphs
    
    while i < len(paras):
        text = paras[i].text.strip()
        
        if text == "Title" and i+1 < len(paras):
            meta_title = paras[i+1].text.strip()
            i += 2
        elif text == "Meta Description" and i+1 < len(paras):
            meta_description = paras[i+1].text.strip()
            i += 2
        elif text == "Keywords":
            i += 2
            break
        else:
            i += 1
    
    h1 = ""  # Blir H2 i HTML
    lead = ""
    sections = []
    current_section = None
    
    while i < len(paras):
        para = paras[i]
        text = para.text.strip()
        style = para.style.name
        
        if not text:
            i += 1
            continue
        
        if "Vanliga frågor och svar" in text:
            break
        
        # Heading 1 i Word → H2 i HTML
        if 'Heading 1' in style and not h1:
            h1 = text
            i += 1
            continue
        
        # Heading 2 i Word → H3 i HTML
        if 'Heading 2' in style:
            if current_section and current_section['content']:
                sections.append(current_section)
            current_section = {'title': text, 'content': []}
            i += 1
            continue
        
        if h1 and not lead and not current_section:
            lead = text
            i += 1
            continue
        
        # Heading 3 i Word → H4 i HTML
        if 'Heading 3' in style and current_section is not None:
            current_section['content'].append({'type': 'h4', 'text': text})
            i += 1
            continue
        
        if current_section is not None:
            numbering_format = get_numbering_format(para, doc)
            
            if numbering_format == 'numbered':
                current_section['content'].append({'type': 'oli', 'text': text})
            elif numbering_format == 'bullet':
                current_section['content'].append({'type': 'li', 'text': text})
            else:
                current_section['content'].append({'type': 'p', 'text': text})
        
        i += 1
    
    if current_section and current_section['content']:
        sections.append(current_section)
    
    return {
        'meta_title': meta_title,
        'meta_description': meta_description,
        'h1': h1,  # Blir H2
        'lead': lead,
        'sections': sections
    }

def data_to_html(data):
    sections_html = []
    for section in data['sections']:
        section_id = slugify(section['title'])
        content_html = []
        in_ul = False
        in_ol = False
        
        for item in section['content']:
            if item['type'] == 'h4':
                if in_ul:
                    content_html.append('</ul>')
                    in_ul = False
                if in_ol:
                    content_html.append('</ol>')
                    in_ol = False
                content_html.append(f'<h4>{item["text"]}</h4>')
            
            elif item['type'] == 'oli':
                if in_ul:
                    content_html.append('</ul>')
                    in_ul = False
                if not in_ol:
                    content_html.append('<ol>')
                    in_ol = True
                content_html.append(f'<li>{item["text"]}</li>')
            
            elif item['type'] == 'li':
                if in_ol:
                    content_html.append('</ol>')
                    in_ol = False
                if not in_ul:
                    content_html.append('<ul>')
                    in_ul = True
                content_html.append(f'<li>{item["text"]}</li>')
            
            else:
                if in_ul:
                    content_html.append('</ul>')
                    in_ul = False
                if in_ol:
                    content_html.append('</ol>')
                    in_ol = False
                content_html.append(f'<p>{item["text"]}</p>')
        
        if in_ul:
            content_html.append('</ul>')
        if in_ol:
            content_html.append('</ol>')
        
        sections_html.append(f'''<section id="{section_id}">
<h3>{section['title']}</h3>
{chr(10).join(content_html)}
</section>
''')
    
    # Bara ren HTML, ingen <head>, <style> eller <body>
    html = f'''<h2>{data['h1']}</h2>
<p class="lead">{data['lead']}</p>
{''.join(sections_html)}'''
    
    return html

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process():
    files = request.files.getlist('files')
    
    if not files or files[0].filename == '':
        return {'error': 'Inga filer uppladdade'}, 400
    
    results = []
    temp_dir = tempfile.mkdtemp()
    
    for file in files:
        if file.filename.endswith('.docx'):
            filename = secure_filename(file.filename)
            filepath = os.path.join(temp_dir, filename)
            file.save(filepath)
            
            try:
                data = parse_docx(filepath)
                html = data_to_html(data)
                
                results.append({
                    'filename': filename,
                    'meta_title': data['meta_title'],
                    'meta_description': data['meta_description'],
                    'html': html
                })
            except Exception as e:
                import traceback
                traceback.print_exc()
                results.append({
                    'filename': filename,
                    'meta_title': '',
                    'meta_description': '',
                    'html': f'FEL: {str(e)}'
                })
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Dokument"
    
    headers = ['Key', 'Name', 'Slug', 'Description', 'MetaTitle', 'MetaDescription']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='top')
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 80
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 60
    
    for result in results:
        ws.append([
            '',
            '',
            '',
            result['html'],
            result['meta_title'],
            result['meta_description']
        ])
    
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    excel_path = os.path.join(temp_dir, 'html_dokument.xlsx')
    wb.save(excel_path)
    
    return send_file(excel_path, as_attachment=True, download_name='html_dokument.xlsx')

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)