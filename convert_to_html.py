import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import unicodedata
import re
from docx import Document

def slugify(text):
    """Konverterar text till URL-vänligt ID"""
    text = unicodedata.normalize('NFKD', text)
    text = text.encode('ascii', 'ignore').decode('ascii')
    text = re.sub(r'[^\w\s-]', '', text.lower())
    text = re.sub(r'[-\s]+', '-', text)
    return text.strip('-')

def is_list_item(para):
    """Kollar om paragrafen är en listpunkt i Word"""
    if 'List' in para.style.name:
        return True
    
    if para._element.pPr is not None:
        numPr = para._element.pPr.numPr
        if numPr is not None:
            return True
    
    return False

def parse_docx(filepath):
    """Läser Word-dokument"""
    doc = Document(filepath)
    
    state = 'skip_metadata'
    h1 = ""
    lead = ""
    sections = []
    current_section = None
    
    for para in doc.paragraphs:
        text = para.text.strip()
        style = para.style.name
        
        if not text:
            continue
        
        if state == 'skip_metadata':
            if text == "Keywords":
                state = 'found_keywords'
            continue
        
        if state == 'found_keywords':
            state = 'content'
            continue
        
        if "Vanliga frågor och svar" in text:
            break
        
        if state == 'content':
            if 'Heading 1' in style and not h1:
                h1 = text
                continue
            
            if 'Heading 2' in style:
                if current_section and current_section['content']:
                    sections.append(current_section)
                current_section = {'title': text, 'content': []}
                continue
            
            if h1 and not lead and not current_section:
                lead = text
                continue
            
            if 'Heading 3' in style and current_section is not None:
                current_section['content'].append({'type': 'h3', 'text': text})
                continue
            
            if current_section is not None:
                if is_list_item(para):
                    current_section['content'].append({'type': 'li', 'text': text})
                else:
                    current_section['content'].append({'type': 'p', 'text': text})
    
    if current_section and current_section['content']:
        sections.append(current_section)
    
    return {
        'h1': h1,
        'lead': lead,
        'sections': sections
    }

def data_to_html(data):
    """Konverterar data till HTML"""
    
    sections_html = []
    for section in data['sections']:
        section_id = slugify(section['title'])
        content_html = []
        in_ul = False
        
        for item in section['content']:
            if item['type'] == 'h3':
                if in_ul:
                    content_html.append('        </ul>')
                    in_ul = False
                content_html.append(f'        <h3>{item["text"]}</h3>')
            elif item['type'] == 'li':
                if not in_ul:
                    content_html.append('        <ul>')
                    in_ul = True
                content_html.append(f'          <li>{item["text"]}</li>')
            else:
                if in_ul:
                    content_html.append('        </ul>')
                    in_ul = False
                content_html.append(f'        <p>{item["text"]}</p>')
        
        if in_ul:
            content_html.append('        </ul>')
        
        sections_html.append(f'''      <section id="{section_id}">
        <h2>{section['title']}</h2>
{chr(10).join(content_html)}
      </section>
''')
    
    html = f'''<!DOCTYPE html>
<html lang="sv">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{data['h1']}</title>
  
  <style>
    body {{ font-family: system-ui, -apple-system, Segoe UI, Roboto, Helvetica, Arial, sans-serif; line-height: 1.6; color: #111; margin: 0; }}
    main {{ max-width: 72ch; margin: 0 auto; padding: 24px; }}
    h1, h2, h3 {{ line-height: 1.25; }}
    ul {{ padding-left: 1.2rem; }}
    .lead {{ font-size: 1.125rem; }}
  </style>
</head>
<body>
  <main>
    <header>
      <h1>{data['h1']}</h1>
      <p class="lead">
        {data['lead']}
      </p>
    </header>

    <article>
{''.join(sections_html)}
    </article>
  </main>
</body>
</html>'''
    
    return html

def process_documents(input_folder, output_file):
    """Bearbetar dokument"""
    
    results = []
    
    for filename in sorted(os.listdir(input_folder)):
        if filename.endswith('.docx'):
            filepath = os.path.join(input_folder, filename)
            
            try:
                data = parse_docx(filepath)
                html = data_to_html(data)
                
                # Räkna sektioner
                num_sections = len(data['sections'])
                section_titles = ', '.join([s['title'] for s in data['sections'][:3]])
                if num_sections > 3:
                    section_titles += f" (+ {num_sections - 3} till)"
                
                results.append({
                    'filename': filename,
                    'h1': data['h1'],
                    'lead_preview': data['lead'][:100] + '...' if len(data['lead']) > 100 else data['lead'],
                    'num_sections': num_sections,
                    'section_titles': section_titles,
                    'html': html
                })
                
                # Spara HTML
                html_filename = filename.replace('.docx', '.html')
                with open(os.path.join(input_folder, html_filename), 'w', encoding='utf-8') as f:
                    f.write(html)
                
                print(f"✓ {filename} → {html_filename}")
                
            except Exception as e:
                print(f"✗ {filename}: {e}")
                import traceback
                traceback.print_exc()
    
    # Skapa Excel med flera kolumner
    wb = Workbook()
    ws = wb.active
    ws.title = "Dokument"
    
    # Headers
    headers = ['Filnamn', 'H1 Titel', 'Lead (förhandsvisning)', 'Antal sektioner', 'Sektioner', 'HTML Kod']
    ws.append(headers)
    
    # Formatera headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='top')
    
    # Kolumnbredder
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 20
    
    # Lägg till data
    for result in results:
        ws.append([
            result['filename'],
            result['h1'],
            result['lead_preview'],
            result['num_sections'],
            result['section_titles'],
            result['html']
        ])
    
    # Formatera alla celler
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    wb.save(output_file)
    
    print(f"\n{'='*60}")
    print(f"✓ Excel-fil skapad: {output_file}")
    print(f"✓ Totalt {len(results)} dokument bearbetade")
    print(f"\nKolumner i Excel:")
    print(f"  - Filnamn")
    print(f"  - H1 Titel")
    print(f"  - Lead (förhandsvisning)")
    print(f"  - Antal sektioner")
    print(f"  - Sektioner (lista)")
    print(f"  - HTML Kod (komplett)")
    print(f"{'='*60}")

if __name__ == "__main__":
    folder = "dokument"
    
    if not os.path.exists(folder):
        os.makedirs(folder)
        print(f"Skapade {folder}/")
        print("Lägg .docx filer i mappen")
    else:
        process_documents(folder, "html_dokument.xlsx")